#!/usr/bin/env python3
"""
Mobtronic LLC Excel API Microservice (Secure Version)
A Flask-based API for capturing website form submissions and storing them in an Excel file.
This version uses secure credential management with .env files.
"""

import os
import smtplib
from datetime import datetime
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from flask import Flask, request, jsonify
from flask_cors import CORS
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv
import getpass

# Load environment variables from .env file
load_dotenv()

app = Flask(__name__)
CORS(app)  # Enable CORS for all routes

# Configuration - Load from .env file
EXCEL_FILE = 'mobtronic_leads.xlsx'
SMTP_SERVER = os.getenv('SMTP_SERVER', 'smtp.office365.com')
SMTP_PORT = int(os.getenv('SMTP_PORT', '587'))
EMAIL_USER = os.getenv('EMAIL_USER', 'support@mobtronic.org')
EMAIL_PASSWORD = os.getenv('EMAIL_PASSWORD', '')  # Will be loaded from .env file
INTERNAL_EMAIL = os.getenv('INTERNAL_EMAIL', 'support@mobtronic.org')

# Excel column headers
EXCEL_HEADERS = [
    'Timestamp', 'Lead ID', 'First Name', 'Last Name', 'Email', 'Company', 
    'Role', 'Phone', 'Topic', 'Notes', 'Consent', 'Source Page', 
    'UTM Source', 'UTM Medium', 'UTM Campaign', 'UTM Term', 'UTM Content'
]

def initialize_excel_file():
    """Initialize the Excel file with headers if it doesn't exist."""
    if not os.path.exists(EXCEL_FILE):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Leads"
        
        # Add headers
        for col_num, header in enumerate(EXCEL_HEADERS, 1):
            worksheet.cell(row=1, column=col_num, value=header)
        
        # Auto-adjust column widths
        for col_num in range(1, len(EXCEL_HEADERS) + 1):
            column_letter = get_column_letter(col_num)
            worksheet.column_dimensions[column_letter].width = 15
        
        workbook.save(EXCEL_FILE)
        print(f"✅ Created Excel file: {EXCEL_FILE}")

def add_lead_to_excel(lead_data):
    """Add a new lead to the Excel file."""
    try:
        workbook = load_workbook(EXCEL_FILE)
        worksheet = workbook.active
        
        # Find the next empty row
        next_row = worksheet.max_row + 1
        
        # Generate lead ID
        lead_id = f"MOB-{datetime.now().strftime('%Y%m%d')}-{next_row:04d}"
        
        # Prepare row data
        row_data = [
            datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            lead_id,
            lead_data.get('firstName', ''),
            lead_data.get('lastName', ''),
            lead_data.get('email', ''),
            lead_data.get('company', ''),
            lead_data.get('role', ''),
            lead_data.get('phone', ''),
            lead_data.get('topic', ''),
            lead_data.get('notes', ''),
            'Yes' if lead_data.get('consent', False) else 'No',
            lead_data.get('sourcePage', ''),
            lead_data.get('utmSource', ''),
            lead_data.get('utmMedium', ''),
            lead_data.get('utmCampaign', ''),
            lead_data.get('utmTerm', ''),
            lead_data.get('utmContent', '')
        ]
        
        # Add data to worksheet
        for col_num, value in enumerate(row_data, 1):
            worksheet.cell(row=next_row, column=col_num, value=value)
        
        workbook.save(EXCEL_FILE)
        print(f"✅ Added lead to Excel: {lead_id}")
        return lead_id
        
    except Exception as e:
        print(f"❌ Error adding lead to Excel: {str(e)}")
        raise

def send_confirmation_email(lead_data, lead_id):
    """Send confirmation email to the user."""
    if not EMAIL_PASSWORD:
        print("⚠️ Email not configured - skipping confirmation email")
        return False
    
    try:
        # Create message
        msg = MIMEMultipart()
        msg['From'] = EMAIL_USER
        msg['To'] = lead_data['email']
        msg['Subject'] = f"Thank you for your interest in Mobtronic LLC - {lead_id}"
        
        # Email body
        topic_services = {
            'infra': 'AI-Ready Infrastructure Audit',
            'fhir': 'FHIR/TEFCA 90-Day Sprint',
            'consolidation': 'Technology Consolidation & Divestiture'
        }
        
        service_name = topic_services.get(lead_data.get('topic', ''), 'Our Services')
        
        body = f"""
Dear {lead_data.get('firstName', '')},

Thank you for your interest in {service_name} from Mobtronic LLC.

We have received your inquiry (Reference: {lead_id}) and our team will review your requirements. 
You can expect to hear from us within 24-48 hours.

In the meantime, you can download our executive briefs and service overviews from your confirmation page.

Best regards,
The Mobtronic Team

---
Mobtronic LLC
Building Infrastructure Behind $4T+ Alternative Investments
Email: support@mobtronic.org
Website: https://mobtronic.org

This email was sent in response to your inquiry on our website. If you did not make this request, please ignore this email.
        """
        
        msg.attach(MIMEText(body, 'plain'))
        
        # Send email
        server = smtplib.SMTP(SMTP_SERVER, SMTP_PORT)
        server.starttls()
        server.login(EMAIL_USER, EMAIL_PASSWORD)
        text = msg.as_string()
        server.sendmail(EMAIL_USER, lead_data['email'], text)
        server.quit()
        
        print(f"✅ Confirmation email sent to {lead_data['email']}")
        return True
        
    except Exception as e:
        print(f"❌ Error sending confirmation email: {str(e)}")
        return False

def send_internal_notification(lead_data, lead_id):
    """Send internal notification email to the team."""
    if not EMAIL_PASSWORD:
        print("⚠️ Email not configured - skipping internal notification")
        return False
    
    try:
        # Create message
        msg = MIMEMultipart()
        msg['From'] = EMAIL_USER
        msg['To'] = INTERNAL_EMAIL
        msg['Subject'] = f"New Lead: {lead_id} - {lead_data.get('company', 'Unknown Company')}"
        
        # Email body
        body = f"""
New lead submission received:

Lead ID: {lead_id}
Timestamp: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}

Contact Information:
- Name: {lead_data.get('firstName', '')} {lead_data.get('lastName', '')}
- Email: {lead_data.get('email', '')}
- Company: {lead_data.get('company', '')}
- Role: {lead_data.get('role', '')}
- Phone: {lead_data.get('phone', '')}

Service Interest: {lead_data.get('topic', '')}
Notes: {lead_data.get('notes', 'None provided')}

Marketing Attribution:
- Source Page: {lead_data.get('sourcePage', '')}
- UTM Source: {lead_data.get('utmSource', '')}
- UTM Medium: {lead_data.get('utmMedium', '')}
- UTM Campaign: {lead_data.get('utmCampaign', '')}

Consent Given: {'Yes' if lead_data.get('consent', False) else 'No'}

---
This notification was generated automatically by the Mobtronic CRM system.
        """
        
        msg.attach(MIMEText(body, 'plain'))
        
        # Send email
        server = smtplib.SMTP(SMTP_SERVER, SMTP_PORT)
        server.starttls()
        server.login(EMAIL_USER, EMAIL_PASSWORD)
        text = msg.as_string()
        server.sendmail(EMAIL_USER, INTERNAL_EMAIL, text)
        server.quit()
        
        print(f"✅ Internal notification sent to {INTERNAL_EMAIL}")
        return True
        
    except Exception as e:
        print(f"❌ Error sending internal notification: {str(e)}")
        return False

@app.route('/submit-lead', methods=['POST'])
def submit_lead():
    """Handle lead submission from the website."""
    try:
        # Get JSON data from request
        lead_data = request.get_json()
        
        if not lead_data:
            return jsonify({'error': 'No data provided'}), 400
        
        # Validate required fields
        required_fields = ['firstName', 'lastName', 'email', 'company']
        for field in required_fields:
            if not lead_data.get(field):
                return jsonify({'error': f'Missing required field: {field}'}), 400
        
        # Add lead to Excel
        lead_id = add_lead_to_excel(lead_data)
        
        # Send emails (non-blocking - don't fail if email fails)
        try:
            send_confirmation_email(lead_data, lead_id)
            send_internal_notification(lead_data, lead_id)
        except Exception as email_error:
            print(f"⚠️ Email sending failed: {str(email_error)}")
        
        # Return success response
        return jsonify({
            'success': True,
            'message': 'Lead submitted successfully',
            'lead_id': lead_id
        }), 200
        
    except Exception as e:
        print(f"❌ Error processing lead submission: {str(e)}")
        return jsonify({'error': 'Internal server error'}), 500

@app.route('/health', methods=['GET'])
def health_check():
    """Health check endpoint."""
    return jsonify({
        'status': 'healthy',
        'service': 'Mobtronic Excel API',
        'timestamp': datetime.now().isoformat(),
        'excel_file': EXCEL_FILE,
        'email_configured': bool(EMAIL_PASSWORD)
    }), 200

@app.route('/', methods=['GET'])
def root():
    """Root endpoint with API information."""
    return jsonify({
        'service': 'Mobtronic LLC Excel API Microservice',
        'version': '1.0.0',
        'description': 'Secure API for capturing website form submissions and storing them in Excel',
        'endpoints': {
            '/submit-lead': 'POST - Submit a new lead',
            '/health': 'GET - Health check',
            '/': 'GET - API information'
        },
        'timestamp': datetime.now().isoformat()
    }), 200

if __name__ == '__main__':
    print("🚀 Starting Mobtronic LLC Excel API Microservice...")
    
    # Initialize Excel file
    initialize_excel_file()
    
    # Check email configuration
    if EMAIL_PASSWORD:
        print("✅ Email configuration detected")
    else:
        print("⚠️ Email not configured - set EMAIL_PASSWORD in .env file")
    
    # Start Flask app
    print("🌐 Starting Flask server on http://localhost:5000")
    app.run(host='0.0.0.0', port=5000, debug=True)
